"""
Export Helper - Xuất dữ liệu ra các định dạng khác nhau
Hỗ trợ: JSON, CSV, Excel, PDF
"""
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Tuple
import logging

from config.settings import AppConfig

logger = logging.getLogger(__name__)


class ExportHelper:
    """Helper class cho các chức năng xuất dữ liệu"""

    # ========== EXPORT READERS ==========

    @staticmethod
    def export_to_json(readers, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách bạn đọc ra file JSON"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"readers_{timestamp}.json"

            data = {
                'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'total_records': len(readers),
                'readers': [reader.to_dict() for reader in readers]
            }

            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True, str(filename)

        except Exception as e:
            logger.error(f"Error exporting JSON: {e}")
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_to_csv(readers, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách bạn đọc ra file CSV"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"readers_{timestamp}.csv"

            with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)

                # Header
                writer.writerow([
                    'ID', 'Họ tên', 'Địa chỉ', 'Điện thoại', 'Email',
                    'Ngày cấp thẻ', 'Ngày hết hạn', 'Trạng thái', 'Điểm uy tín'
                ])

                # Data
                for reader in readers:
                    writer.writerow([
                        reader.reader_id or '',
                        reader.full_name or '',
                        reader.address or '',
                        reader.phone or '',
                        reader.email or '',
                        reader.card_start or '',
                        reader.card_end or '',
                        reader.status or '',
                        reader.reputation_score or 0
                    ])

            return True, str(filename)

        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_to_excel(readers, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách bạn đọc ra file Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"readers_{timestamp}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách Bạn đọc"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Headers
            headers = [
                'ID', 'Họ tên', 'Địa chỉ', 'Điện thoại', 'Email',
                'Ngày cấp thẻ', 'Ngày hết hạn', 'Trạng thái', 'Điểm uy tín'
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            for row, reader in enumerate(readers, start=2):
                ws.cell(row=row, column=1, value=reader.reader_id or '')
                ws.cell(row=row, column=2, value=reader.full_name or '')
                ws.cell(row=row, column=3, value=reader.address or '')
                ws.cell(row=row, column=4, value=reader.phone or '')
                ws.cell(row=row, column=5, value=reader.email or '')
                ws.cell(row=row, column=6, value=reader.card_start or '')
                ws.cell(row=row, column=7, value=reader.card_end or '')
                ws.cell(row=row, column=8, value=reader.status or '')
                ws.cell(row=row, column=9, value=reader.reputation_score or 0)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            return True, str(filename)

        except ImportError:
            return False, "Chưa cài đặt thư viện openpyxl. Chạy: pip install openpyxl"
        except Exception as e:
            logger.error(f"Error exporting Excel: {e}")
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_to_pdf(readers, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách bạn đọc ra file PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"readers_{timestamp}.pdf"

            # Create PDF
            doc = SimpleDocTemplate(
                str(filename),
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=18
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=30,
                alignment=1
            )

            # Title
            title = Paragraph("DANH SÁCH BẠN ĐỌC", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2 * inch))

            # Info
            info_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Tổng số: {len(readers)} bạn đọc"
            info = Paragraph(info_text, styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 0.3 * inch))

            # Table data
            data = [['ID', 'Họ tên', 'Điện thoại', 'Email', 'Ngày cấp', 'Ngày HH', 'Trạng thái', 'Điểm']]

            for reader in readers:
                data.append([
                    str(reader.reader_id or ''),
                    (reader.full_name or '')[:30],
                    reader.phone or 'N/A',
                    (reader.email or 'N/A')[:25],
                    reader.card_start or 'N/A',
                    reader.card_end or 'N/A',
                    reader.status or 'N/A',
                    str(reader.reputation_score or 0)
                ])

            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)
            doc.build(elements)

            return True, str(filename)

        except ImportError:
            return False, "Chưa cài đặt thư viện reportlab. Chạy: pip install reportlab"
        except Exception as e:
            logger.error(f"Error exporting PDF: {e}")
            return False, f"Lỗi: {str(e)}"

    # ========== EXPORT BOOKS ==========

    @staticmethod
    def export_books_to_json(books, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách sách ra file JSON"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"books_{timestamp}.json"

            data = {
                'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'total_records': len(books),
                'books': [book.to_dict() for book in books]
            }

            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            return True, str(filename)

        except Exception as e:
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_books_to_csv(books, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách sách ra file CSV"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"books_{timestamp}.csv"

            with open(filename, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)

                # Header
                writer.writerow([
                    'ID', 'Tựa sách', 'Tác giả', 'Thể loại', 'NXB',
                    'Năm XB', 'ISBN', 'Barcode', 'Giá',
                    'Tổng SL', 'Còn', 'Mô tả'
                ])

                # Data
                for book in books:
                    writer.writerow([
                        book.book_id or '',
                        book.title or '',
                        book.author_name or '',
                        book.category_name or '',
                        book.publisher_name or '',
                        book.publish_year or '',
                        book.isbn or '',
                        book.barcode or '',
                        book.price or '',
                        book.total_quantity or 0,
                        book.available_quantity or 0,
                        book.description or ''
                    ])

            return True, str(filename)

        except Exception as e:
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_books_to_excel(books, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách sách ra file Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"books_{timestamp}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách Sách"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Headers
            headers = [
                'ID', 'Tựa sách', 'Tác giả', 'Thể loại', 'NXB',
                'Năm XB', 'ISBN', 'Barcode', 'Giá (VNĐ)',
                'Tổng SL', 'Còn', 'Trạng thái', 'Mô tả'
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            for row, book in enumerate(books, start=2):
                ws.cell(row=row, column=1, value=book.book_id or '')
                ws.cell(row=row, column=2, value=book.title or '')
                ws.cell(row=row, column=3, value=book.author_name or '')
                ws.cell(row=row, column=4, value=book.category_name or '')
                ws.cell(row=row, column=5, value=book.publisher_name or '')
                ws.cell(row=row, column=6, value=book.publish_year or '')
                ws.cell(row=row, column=7, value=book.isbn or '')
                ws.cell(row=row, column=8, value=book.barcode or '')
                ws.cell(row=row, column=9, value=book.price or 0)
                ws.cell(row=row, column=10, value=book.total_quantity or 0)
                ws.cell(row=row, column=11, value=book.available_quantity or 0)
                ws.cell(row=row, column=12, value=book.get_stock_status())
                ws.cell(row=row, column=13, value=book.description or '')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filename)
            return True, str(filename)

        except ImportError:
            return False, "Chưa cài đặt thư viện openpyxl. Chạy: pip install openpyxl"
        except Exception as e:
            return False, f"Lỗi: {str(e)}"

    @staticmethod
    def export_books_to_pdf(books, filename: str = None) -> Tuple[bool, str]:
        """Xuất danh sách sách ra file PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = AppConfig.EXPORT_DIR / f"books_{timestamp}.pdf"

            # Create PDF
            doc = SimpleDocTemplate(
                str(filename),
                pagesize=landscape(A4),
                rightMargin=30,
                leftMargin=30,
                topMargin=30,
                bottomMargin=18
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1976D2'),
                spaceAfter=30,
                alignment=1
            )

            # Title
            title = Paragraph("DANH SÁCH SÁCH", title_style)
            elements.append(title)
            elements.append(Spacer(1, 0.2 * inch))

            # Info
            info_text = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Tổng số: {len(books)} sách"
            info = Paragraph(info_text, styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 0.3 * inch))

            # Table data
            data = [['ID', 'Tựa sách', 'Tác giả', 'Thể loại', 'NXB', 'Năm', 'ISBN', 'Giá', 'Tồn kho']]

            for book in books:
                price_str = f"{book.price:,.0f}" if book.price else "0"
                data.append([
                    str(book.book_id or ''),
                    (book.title or '')[:30] + '...' if book.title and len(book.title) > 30 else (book.title or ''),
                    (book.author_name or '')[:20],
                    (book.category_name or '')[:15],
                    (book.publisher_name or '')[:20],
                    str(book.publish_year or ''),
                    (book.isbn or '')[:15],
                    price_str,
                    f"{book.available_quantity}/{book.total_quantity}"
                ])

            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)
            doc.build(elements)

            return True, str(filename)

        except ImportError:
            return False, "Chưa cài đặt thư viện reportlab. Chạy: pip install reportlab"
        except Exception as e:
            return False, f"Lỗi: {str(e)}"